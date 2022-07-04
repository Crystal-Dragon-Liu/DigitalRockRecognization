from bdb import GENERATOR_AND_COROUTINE_FLAGS
import enum
import imp
from logging import root
import os
from pickletools import int4
from unittest.mock import patch
import cv2
from tqdm import tqdm
from PIL import Image
import numpy as np
import matplotlib.pyplot as plt
import sys
import operator
sys.path.append('..')
import json
import imgviz.label as il
root_path = '/Users/crystal-dragon-lyb/PycharmProjects/DigitalRockSrc/dataset_rock/'
raw_data_path = root_path + 'raw_data/'
raw_mask_path = root_path + 'raw_mask/'
img_save_path = root_path + 'image/'
mask_save_path = root_path + 'mask/'
json_save_path = root_path + 'meta_data/'
json_data_path = json_save_path + 'rgb_meta.json'
excel_save_path = json_data_path + 'color_map.xlsx'
new_dataset_path = '/Users/crystal-dragon-lyb/PycharmProjects/DigitalRockSrc/'
mask_postfix_path = '/VOC2007/SegmentationClass/'
IMG_INDEX = 0
MASK_INDEX = 0
SPLIT_FILES = False
COUNT_RGB = False
RECORD_COLOR = False
DRAW_COUNT_PLOT = False
GEN_NEW_MASK = True
# some functions
# ===================================================================================================================
def get_img_to_array(img_path: str):
    img = cv2.imread(img_path)
    return img

def get_rgb_list(rgb_list_str: str, split_str = '_'):
    temp = rgb_list_str.split(split_str)
    result = [int(i) for i in temp]
    return result


def split_and_save_rock_data(img_name: str, img_path: str, save_path: str = '', size: tuple=(512, 512)):
    img = cv2.imread(img_path)
    print("size: ", img.shape)
    height          =   img.shape[0]
    width           =   img.shape[1]
    row_batch       = height // size[0]
    column_batch    = width // size[1]
    i = 0
    print("batchSIze:", row_batch, column_batch)
    for row in tqdm(range(row_batch)):
        for column in range(column_batch):
            row_start       =   row*size[0]
            row_end         =   row*size[0]+ size[0]
            column_start    =   column*size[1]
            column_end      =   column* size[1] + size[1]
            data_block = img[row_start: row_end, column_start: column_end]
            cv2.imwrite(save_path + img_name + "_"+ str(i) + '.png', img[row_start: row_end, column_start: column_end], [cv2.IMWRITE_PNG_COMPRESSION, 0])
            i = i + 1
    

def get_unique_rgb_list(img_path: str):
    na = get_img_to_array(img_path)
    na = cv2.cvtColor(na, cv2.COLOR_BGR2RGB)
    colours, counts = np.unique(na.reshape(-1,3), axis=0, return_counts=1)
    return colours, counts

def rgb_matched(src_rgb: list, tar_rgb: list)-> bool:
    if len(src_rgb) != len(tar_rgb):
        return False
    for i, ele in enumerate(src_rgb):
        if ele != tar_rgb[i]:
            return False
    return True

def rgb2str(color: list)->str:
    result = str(color[0])
    color_temp = color[1:]
    for c in color_temp:
        result = result + "_" + str(c)
    return result

def get_rgb_str_list_from_jsondata(json_data: list):
    result = []
    for data in json_data:
        result.append(data['rgb'])
    return result

def get_counts_list_from_jsondata(json_data: list):
    result = []
    for data in json_data:
        result.append(int(data['counts']))
    return result

def update_json_data(json_data: list, key: str,new_value: str):
    for data in json_data:
        if data['rgb'] == key:
            data['counts'] = str(int(data['counts']) + int(new_value))
    return json_data

def RGB_to_Hex(tmp, split_char='_', reserve_symbol=True):
    rgb = tmp.split(split_char)#将RGB格式划分开来
    result = '#'
    for i in rgb:
        num = int(i)#将str转int
        #将R、G、B分别转化为16进制拼接转换并大写
        result += str(hex(num))[-2:].replace('x','0').upper()
    if reserve_symbol == True:
        result = result[1:]
    return result

def upload_rgb_json(img_path: str, json_path: str):
    print('processing img -> ', img_path)
    colours, counts = get_unique_rgb_list(img_path)
    try: 
        rgb_meta_temp = json.load(open(json_path))
    except: 
        rgb_meta_temp = []
    if not os.path.exists(json_path) and len(rgb_meta_temp) == 0: 
        print('creating a new json file...')
        temp = open(json_path, 'w')
        temp.close()
    # if json file is empty, then create a empty list.

    old_rgb_list = get_rgb_str_list_from_jsondata(rgb_meta_temp)
    with open(json_path, 'w') as outfile:
        for i, color in enumerate(tqdm(colours)):
            rgb_str = rgb2str(color)
            count_str = str(counts[i])
            # if this color has not been recorded in json file, 
            # upload this color and the corresponding counts
            if len(old_rgb_list) == 0 or rgb_str not in old_rgb_list:
                rgb_meta_temp.append({'rgb': rgb_str, 'counts': count_str})
            else:
                # update json data.
                rgb_meta_temp = update_json_data(rgb_meta_temp, rgb_str, count_str)
        json.dump(rgb_meta_temp, outfile)
    print('metadata < ', img_path ,' > uploaded.')


"""
 Plot - adjust axes
"""
def adjust_axes(r, t, fig, axes):
    # get text width for re-scaling
    bb = t.get_window_extent(renderer=r)
    text_width_inches = bb.width / fig.dpi
    # get axis width in inches
    current_fig_width = fig.get_figwidth()
    new_fig_width = current_fig_width + text_width_inches
    propotion = new_fig_width / current_fig_width
    # get axis limit
    x_lim = axes.get_xlim()
    axes.set_xlim([x_lim[0], x_lim[1]*propotion])

"""
 Draw plot using Matplotlib
"""
def draw_plot_func(dictionary, n_classes, window_title, plot_title, x_label, output_path, to_show, plot_color, true_p_bar):
    # sort the dictionary by decreasing value, into a list of tuples
    sorted_dic_by_value = sorted(dictionary.items(), key=operator.itemgetter(1))
    # unpacking the list of tuples into two lists
    sorted_keys, sorted_values = zip(*sorted_dic_by_value)
    # 
    if true_p_bar != "":
        """
         Special case to draw in:
            - green -> TP: True Positives (object detected and matches ground-truth)
            - red -> FP: False Positives (object detected but does not match ground-truth)
            - orange -> FN: False Negatives (object not detected but present in the ground-truth)
        """
        fp_sorted = []
        tp_sorted = []
        for key in sorted_keys:
            fp_sorted.append(dictionary[key] - true_p_bar[key])
            tp_sorted.append(true_p_bar[key])
        plt.barh(range(n_classes), fp_sorted, align='center', color='crimson', label='False Positive')
        plt.barh(range(n_classes), tp_sorted, align='center', color='forestgreen', label='True Positive', left=fp_sorted)
        # add legend
        plt.legend(loc='lower right')
        """
         Write number on side of bar
        """
        fig = plt.gcf() # gcf - get current figure
        axes = plt.gca()
        r = fig.canvas.get_renderer()
        for i, val in enumerate(sorted_values):
            fp_val = fp_sorted[i]
            tp_val = tp_sorted[i]
            fp_str_val = " " + str(fp_val)
            tp_str_val = fp_str_val + " " + str(tp_val)
            # trick to paint multicolor with offset:
            # first paint everything and then repaint the first number
            t = plt.text(val, i, tp_str_val, color='forestgreen', va='center', fontweight='bold')
            plt.text(val, i, fp_str_val, color='crimson', va='center', fontweight='bold')
            if i == (len(sorted_values)-1): # largest bar
                adjust_axes(r, t, fig, axes)
    else:
        plt.barh(range(n_classes), sorted_values, color=plot_color)
        """
         Write number on side of bar
        """
        fig = plt.gcf() # gcf - get current figure
        axes = plt.gca()
        r = fig.canvas.get_renderer()
        for i, val in enumerate(sorted_values):
            str_val = " " + str(val) # add a space before
            if val < 1.0:
                str_val = " {0:.2f}".format(val)
            t = plt.text(val, i, str_val, color=plot_color, va='center', fontweight='bold')
            # re-set axes to show number inside the figure
            if i == (len(sorted_values)-1): # largest bar
                adjust_axes(r, t, fig, axes)
    # set window title
    fig.canvas.set_window_title(window_title)
    # write classes in y axis
    tick_font_size = 12
    plt.yticks(range(n_classes), sorted_keys, fontsize=tick_font_size)
    """
     Re-scale height accordingly
    """
    init_height = fig.get_figheight()
    # comput the matrix height in points and inches
    dpi = fig.dpi
    height_pt = n_classes * (tick_font_size * 1.4) # 1.4 (some spacing)
    height_in = height_pt / dpi
    # compute the required figure height 
    top_margin = 0.15 # in percentage of the figure height
    bottom_margin = 0.05 # in percentage of the figure height
    figure_height = height_in / (1 - top_margin - bottom_margin)
    # set new height
    if figure_height > init_height:
        fig.set_figheight(figure_height)

    # set plot title
    plt.title(plot_title, fontsize=14)
    # set axis titles
    # plt.xlabel('classes')
    plt.xlabel(x_label, fontsize='large')
    # adjust size of window
    fig.tight_layout()
    # save the plot
    fig.savefig(output_path)
    # show image
    if to_show:
        plt.show()
    # close the plot
    plt.close()

def detect_color_from_rgb_list(rgb, rgb_list: list) -> int:
    for index, rgb_str in enumerate(rgb_list):
        rgb_temp = get_rgb_list(rgb_str)
        if rgb[0] == rgb_temp[0] and rgb[1] == rgb_temp[1] and rgb[2] == rgb_temp[2]:
            return index + 1
    return 0

def detect_all_color_from_rgb_list(rgb, rgb_list: list)->int:
    for index, rgb_str in enumerate(rgb_list):
        rgb_temp = get_rgb_list(rgb_str)
        if rgb[0] == rgb_temp[0] and rgb[1] == rgb_temp[1] and rgb[2] == rgb_temp[2]:
            return 1
    return 0

def save_lbl(mask_path: str, lbl):
    if lbl.min() >= -1 and lbl.max() < 255:
        lbl_pil = Image.fromarray(lbl.astype(np.uint8), mode='P')
        colormap = il.label_colormap()
        lbl_pil.putpalette(colormap.flatten())
        lbl_pil.save(mask_path)
    else:
        raise ValueError(
            '[%s] Cannot save the pixel-wise class label as PNG. '
            'Please consider using the .npy format.' % mask_path
        )
    pass

def generate_new_label(old_mask_path: str, new_mask_path: str, rgb_list: list, detect_func = detect_color_from_rgb_list):
    mask_name_list = os.listdir(old_mask_path)
    mask_name_list.remove('.DS_Store')
    for mask_name in tqdm(mask_name_list):
        mask_data = get_img_to_array(old_mask_path + mask_name)
        # print(mask_name)
        mask_data = cv2.cvtColor(mask_data, cv2.COLOR_BGR2RGB)
        mask_height = mask_data.shape[0]
        mask_width = mask_data.shape[1]
        new_lbl = np.zeros(shape=[mask_height, mask_width], dtype=np.int32)
        for i in range(mask_height):
            for j in range(mask_width):
                # TODO detect the color index.
                value = detect_func(mask_data[i,j,:], rgb_list)
                new_lbl[i][j] = value
        new_mask_name_path = new_mask_path + mask_name.split('.')[0] + '.png'
        save_lbl(new_mask_name_path, new_lbl)
# ===================================================================================================================


if SPLIT_FILES == True:
    raw_data_list = os.listdir(raw_data_path)
    raw_mask_list = os.listdir(raw_mask_path)
    raw_data_list.remove('.DS_Store')
    # raw_mask_list.remove('.DS_Store')
    for data_path in raw_data_list:
        img_name = str.split(data_path, '.')[0]
        print(raw_data_path+ data_path)
        split_and_save_rock_data(img_name, raw_data_path + data_path, img_save_path, (128, 128))
        split_and_save_rock_data(img_name, raw_mask_path + data_path, mask_save_path, (128, 128))

# count how many colors there are.
mask_path_list = os.listdir(mask_save_path)
mask_path_list.remove('.DS_Store')

if COUNT_RGB == True:
    for mask_path in mask_path_list:
        upload_rgb_json(mask_save_path + mask_path, json_data_path)

meta_data = json.load(open(json_data_path))
rgb_list = get_rgb_str_list_from_jsondata(meta_data)
counts_list = get_counts_list_from_jsondata(meta_data)


if RECORD_COLOR:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import colors, PatternFill

    wb = Workbook()
    ws = wb.create_sheet('color_map', 0)
    ws.sheet_properties.tabColor = 'ff72BA'
    print("writing colormap to excel file........")
    for index, rgb in enumerate(tqdm(rgb_list)):
        cell = ws.cell(row=index+1, column=1)
        c = RGB_to_Hex(rgb)
        cell.fill = PatternFill(fill_type='solid', start_color=c, end_color=c)# fill background.
        cell.value = rgb
        cell2 = ws.cell(row=index+1, column=2)
        cell2.value = counts_list[index]
    wb.save(excel_save_path)
    wb.close()

color_map = {}
for index, rgb in enumerate(rgb_list):
    # if counts_list[index] == 706565 or counts_list[index] == 815426: #钠长石+石英
    # if counts_list[index] == 189991: # 孔隙
    if counts_list[index] == 187357  or counts_list[index] == 239202 or counts_list[index] == 280688 or counts_list[index] == 48881: # 钾长石+绿泥石+方解石
        color_map[rgb] = counts_list[index]
n_classes = len(color_map.keys())

if DRAW_COUNT_PLOT:
    window_title = "rock counts"
    plot_title = window_title
    x_label = "counts"
    output_path = json_save_path+  "rock_counts.png"
    to_show = True
    plot_color = 'royalblue'
    draw_plot_func(
            color_map,
            n_classes,
            window_title,
            plot_title,
            x_label,
            output_path,
            to_show,
            plot_color,
            ""
            )
# generate new_data.
if GEN_NEW_MASK:
    print('n_classes: ', n_classes)
    
    new_mask_save_path = new_dataset_path + 'VOCdevkit_ckc/' + mask_postfix_path
    if not os.path.exists(new_mask_save_path):
        os.makedirs(new_mask_save_path)
    generate_new_label(mask_save_path, new_mask_save_path, color_map.keys(), detect_func=detect_all_color_from_rgb_list)




